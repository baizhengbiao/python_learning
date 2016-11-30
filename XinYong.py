#!/usr/bin/env python  
# -*- coding:utf-8 -*-  

__author__ = 'Knife'
import shutil, os, string
import xdrlib ,sys
import xlrd
import datetime
import time
from  openpyxl  import  Workbook
from  openpyxl.writer.excel  import  ExcelWriter  
from  openpyxl.utils  import  get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors

import sys
print(sys.version)
#import xlutils
#from xlutils.copy import copy

#当天日期 test
today = datetime.date.today()
dayOfWeek = datetime.date.today().weekday() # 0-6 周一到周二
 
    
#=======================================#写文件
wb = Workbook()
#新建一个excelWriter   
##ew = ExcelWriter(workbook = wb)

#设置文件输出路径与名称   
dest_filename = r'信用汇总.xlsx' 
#第一个sheet是ws
ws = wb.create_sheet()
ws.title = "信用汇总"  
#ws1.
curpath = sys.path
rootdir = curpath[0] + "/分公司信用汇总"                                # 指明被遍历的文件夹
 
indexRow = 1
Count = 0

filePaths = os.walk(rootdir)
print(filePaths)
for dirnames in filePaths:
    
    fenGongSi= dirnames[0]
 
    fenGongSiName = fenGongSi.split("/")
    for dirnameMY in  dirnames[2]:  #输出文件夹信息
        Count += 1
        workbook1= xlrd.open_workbook(dirnames[0]+"/"+dirnameMY)
        table = workbook1.sheet_by_name("信用等级") #通过索引顺序获取
        nrows = table.nrows 
        ncols = table.ncols
        col = get_column_letter(1)
        ws.cell(column = 1,row = indexRow,value =fenGongSiName[-1])
        for i in range(nrows ):  #找类别行
           for j in range(ncols ):  #找类别列                
                 strig2 = table.cell(i,j).value
                # print(strig2)
                 if "客户简称" ==strig2 :
                    ws.cell(column=2, row=indexRow,value = table.cell(i,j+1).value)
                    print(fenGongSiName[-1]+'  '+ strig2+':'+table.cell(i,j+1).value)
                 if "客户信用得分" ==strig2 :
                    ws.cell(column=3, row=indexRow, value=table.cell(i, j + 1).value)
                 if "客户信用等级" ==strig2 :
                    ws.cell(column=4, row=indexRow, value=table.cell(i, j + 1).value)
                 if "客户类型" ==strig2 :
                    ws.cell(column=5, row=indexRow, value=table.cell(i, j + 1).value)
                 if "企业性质" ==strig2 :
                    ws.cell(column=6, row=indexRow, value=table.cell(i, j + 1).value)
                 if "资产状况" ==strig2 :
                    ws.cell(column=7, row=indexRow, value=table.cell(i, j + 1).value)
                 if "门店状况" ==strig2 :
                    ws.cell(column=8, row=indexRow, value=table.cell(i, j + 1).value)
                 if "入行时间" ==strig2 :
                    ws.cell(column=9, row=indexRow, value=table.cell(i, j + 1).value)
                 if "实际控制人情况" ==strig2 :
                    ws.cell(column=10, row=indexRow, value=table.cell(i, j + 1).value)
                 if "月均销量" ==strig2 :
                    ws.cell(column=11, row=indexRow, value=table.cell(i, j + 1).value)
                 if "财务制度建设" ==strig2 :
                    ws.cell(column=12, row=indexRow, value=table.cell(i, j + 1).value)
                 if "销售渠道" ==strig2 :
                    ws.cell(column=13, row=indexRow, value=table.cell(i, j + 1).value)
                 if "信用资料收集情况" ==strig2 :
                    ws.cell(column=14, row=indexRow, value=table.cell(i, j + 1).value)
                 if "担保抵押情况" ==strig2 :
                    ws.cell(column=15, row=indexRow, value=table.cell(i, j + 1).value)
                 if "逾期次数" ==strig2 :
                    ws.cell(column=16, row=indexRow, value=table.cell(i, j + 1).value)
                 if "逾期天数" ==strig2 :
                    ws.cell(column=17, row=indexRow, value=table.cell(i, j + 1).value)
                 if "返利保障执行" ==strig2 :
                    ws.cell(column=18, row=indexRow, value=table.cell(i, j + 1).value)
                 if "自行扣除折扣" ==strig2 :
                    ws.cell(column=19, row=indexRow, value=table.cell(i, j + 1).value)
                 if "对账情况" ==strig2 :
                    ws.cell(column=20, row=indexRow, value=table.cell(i, j + 1).value)
                 if "退票记录" ==strig2 :
                    ws.cell(column=21, row=indexRow, value=table.cell(i, j + 1).value)
                 if "不正常交易情况" ==strig2 :
                    ws.cell(column=22, row=indexRow, value=table.cell(i, j + 1).value)
                 if "工商部门信用记录" ==strig2 :
                    ws.cell(column=23, row=indexRow, value=table.cell(i, j + 1).value)
                 if "税务部门信用记录" ==strig2 :
                    ws.cell(column=24, row=indexRow, value=table.cell(i, j + 1).value)
                 if "法院诉讼记录" ==strig2 :
                    ws.cell(column=25, row=indexRow, value=table.cell(i, j + 1).value)
        indexRow +=1
        #workbook1.
     
print('处理文件个数：%d'%(Count))
wb.save(filename = dest_filename)
 
